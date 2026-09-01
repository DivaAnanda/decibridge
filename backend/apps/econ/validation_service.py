"""Import a validation workbook, apply it, run the engines, and score it (R6).

Produces a PASS/FAIL report comparing computed results against the workbook's
expected values within stated tolerances — the "validation report" the lecturer
asked for. Also surfaces structural validation issues (bad ranges, duplicates,
missing data source) found while importing.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation

from django.db import transaction

from openpyxl import load_workbook

from .lecturer_workbook import is_lecturer_workbook, parse_lecturer_workbook
from .models import Alternative, EconomicModel, EconomicParameter
from .service import IncompleteModelError, run_bia, run_deterministic, run_psa
from .validation_workbook import parse_workbook

_UNIT_INTERVAL = {"probability", "utility", "disutility"}


def _to_decimal(value) -> Decimal | None:
    if value in (None, "", "None"):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _to_int(value) -> int | None:
    if value in (None, "", "None"):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _validate_params(rows: list[dict], issues: list[str]) -> list[dict]:
    seen: set[tuple] = set()
    specs: list[dict] = []
    for p in rows:
        key = str(p.get("key") or "").strip()
        alt = str(p.get("alternative") or "").strip()
        if not key or not alt:
            issues.append("Baris parameter tanpa key/alternative dilewati.")
            continue
        year_index = _to_int(p.get("year_index"))
        value = _to_decimal(p.get("value"))
        ptype = str(p.get("param_type") or "").strip()
        dstatus = str(p.get("data_status") or "").strip()

        if value is None:
            issues.append(f"{key}/{alt}: nilai tidak valid.")
            continue
        if not dstatus:
            issues.append(f"{key}/{alt}: status data (sumber) kosong.")
        if ptype in _UNIT_INTERVAL and not (Decimal("0") <= value <= Decimal("1")):
            issues.append(f"{key}/{alt}: {ptype} di luar rentang 0–1.")
        if ptype == "cost" and value < 0:
            issues.append(f"{key}/{alt}: biaya negatif.")

        dup = (key, alt, year_index)
        if dup in seen:
            issues.append(f"Parameter duplikat: {key}/{alt}/tahun {year_index}.")
            continue
        seen.add(dup)

        specs.append({
            "key": key, "alternative": alt, "year_index": year_index, "value": value,
            "param_type": ptype or "cost", "unit": str(p.get("unit") or "").strip(),
            "data_status": dstatus or "assumption",
            "distribution": str(p.get("distribution") or "fixed").strip() or "fixed",
            "dist_param1": _to_decimal(p.get("dist_param1")),
            "dist_param2": _to_decimal(p.get("dist_param2")),
        })
    return specs


def _apply(case, scalars: dict, specs: list[dict], user) -> EconomicModel:
    model = EconomicModel.objects.filter(case=case).first()
    if model is None:
        model = EconomicModel(case=case, created_by=user)
    model.horizon_years = _to_int(scalars.get("horizon_years")) or 1
    model.cost_discount_rate = _to_decimal(scalars.get("cost_discount_rate")) or Decimal("0")
    model.outcome_discount_rate = _to_decimal(scalars.get("outcome_discount_rate")) or Decimal("0")
    wtp = _to_decimal(scalars.get("wtp_threshold"))
    if wtp is not None:
        model.wtp_threshold = wtp
    model.annual_budget_baseline = _to_decimal(scalars.get("annual_budget_baseline"))
    model.last_edited_by = user
    model.save()

    for s in specs:
        EconomicParameter.objects.update_or_create(
            economic_model=model,
            key=s["key"],
            alternative=s["alternative"],
            year_index=s["year_index"],
            defaults={
                "value": s["value"], "param_type": s["param_type"], "unit": s["unit"],
                "data_status": s["data_status"], "distribution": s["distribution"],
                "dist_param1": s["dist_param1"], "dist_param2": s["dist_param2"],
                "created_by": user, "last_edited_by": user,
            },
        )
    return model


def _check(metric: str, expected, tolerance, actual) -> dict:
    exp = _to_decimal(expected)
    tol = _to_decimal(tolerance) or Decimal("0")
    act = _to_decimal(actual)
    if exp is None or act is None:
        return {"metric": metric, "expected": str(expected), "actual": None if act is None else str(act),
                "diff": None, "tolerance": str(tolerance), "pass": False}
    diff = abs(act - exp)
    return {"metric": metric, "expected": str(exp), "actual": str(act),
            "diff": str(diff), "tolerance": str(tol), "pass": diff <= tol}


def _parse_any_format(file) -> dict:
    """Accept either the lecturer's workbook or DeciBridge's own template."""
    try:
        probe = load_workbook(file, read_only=True, data_only=True)
        sheetnames = probe.sheetnames
    except Exception:
        sheetnames = []
    if hasattr(file, "seek"):
        file.seek(0)
    if is_lecturer_workbook(sheetnames):
        return parse_lecturer_workbook(file)
    return parse_workbook(file)


@transaction.atomic
def import_and_validate(case, file, user) -> dict:
    parsed = _parse_any_format(file)
    if parsed.get("missing_sheets"):
        return {"status": "FAIL", "issues": [f"Sheet wajib hilang: {', '.join(parsed['missing_sheets'])}"],
                "checks": []}

    issues: list[str] = []
    meta_case_id = parsed["case_meta"].get("case_id")
    if meta_case_id and meta_case_id != case.case_id:
        issues.append(f"case_id workbook ({meta_case_id}) tidak konsisten dengan kasus ({case.case_id}).")

    specs = _validate_params(parsed["params"], issues)
    model = _apply(case, parsed["model_scalars"], specs, user)

    checks: list[dict] = []
    try:
        det = run_deterministic(model, computed_by=user)
    except IncompleteModelError as exc:
        issues.append(f"Deterministik tidak dapat dihitung: {exc}")
        return {"status": "FAIL", "issues": issues, "checks": checks}

    actual = {
        "total_cost_intervention": det.total_cost_intervention,
        "total_cost_comparator": det.total_cost_comparator,
        "total_qaly_intervention": det.total_qaly_intervention,
        "total_qaly_comparator": det.total_qaly_comparator,
        "incremental_cost": det.incremental_cost,
        "incremental_qaly": det.incremental_qaly,
        "icer": det.icer,
        "inb": det.inb,
    }
    # BIA scenario metrics (QC09-QC11) — only run the BIA if they're expected.
    if any(str(e["metric"]).startswith("bia_net_") for e in parsed["expected_deterministic"]):
        try:
            bia = run_bia(model, computed_by=user)
            for row in bia.scenarios:
                actual[f"bia_net_{row['label']}"] = Decimal(row["net_budget_impact"])
        except IncompleteModelError as exc:
            issues.append(f"BIA tidak dapat dihitung: {exc}")

    for e in parsed["expected_deterministic"]:
        checks.append(_check(e["metric"], e["expected"], e["tolerance"], actual.get(e["metric"])))

    # PSA validation (fixed seed + n from the workbook → reproducible).
    for e in parsed["expected_psa"]:
        n = _to_int(e.get("n_simulations")) or 1000
        seed = _to_int(e.get("seed")) or 42
        try:
            psa = run_psa(model, computed_by=user, n_simulations=n, seed=seed)
            checks.append(_check(e["metric"], e["expected"], e["tolerance"], psa.prob_cost_effective_base))
        except IncompleteModelError:
            issues.append("PSA tidak dapat dihitung (parameter kurang).")

    all_pass = bool(checks) and all(c["pass"] for c in checks) and not issues
    return {
        "status": "PASS" if all_pass else "FAIL",
        "issues": issues,
        "checks": checks,
        "econ_result_id": det.pk,
    }
