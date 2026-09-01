"""Parser for the lecturer's own validation workbook format.

Source file: `DeciBridge_Economic_Validation_Model_ACEI_Dual.xlsx`, sheets
`01_INPUTS`, `04_PSA_INPUTS`, `09_DATA_MAP_QC` (plus `02_DETERMINISTIC` /
`03_BIA` which hold the expected values referenced by the QC sheet).

We support BOTH workbook shapes:
  * this module  — the lecturer's real file, uploaded as-is;
  * `validation_workbook.py` — DeciBridge's own template.

`validation_service` sniffs the sheet names and dispatches accordingly, so the
QC report is produced either way.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .models import Alternative, DataStatus, ParamKey, ParamType

LECTURER_SHEETS = {"01_INPUTS", "02_DETERMINISTIC", "03_BIA"}

# 01_INPUTS "Parameter" label -> (ParamKey, Alternative, ParamType)
# Labels are matched case-insensitively after whitespace/dash normalisation.
_PARAM_MAP: dict[str, tuple[str, str, str]] = {
    "arni rehospitalization probability": (ParamKey.EVENT_PROBABILITY, Alternative.INTERVENTION, ParamType.PROBABILITY),
    "acei rehospitalization probability": (ParamKey.EVENT_PROBABILITY, Alternative.COMPARATOR, ParamType.PROBABILITY),
    "arni annual medicine cost": (ParamKey.DRUG_COST, Alternative.INTERVENTION, ParamType.COST),
    "acei annual medicine cost": (ParamKey.DRUG_COST, Alternative.COMPARATOR, ParamType.COST),
    "arni median los": (ParamKey.MEDIAN_LOS, Alternative.INTERVENTION, ParamType.COUNT),
    "acei median los": (ParamKey.MEDIAN_LOS, Alternative.COMPARATOR, ParamType.COUNT),
    "arni monitoring cost": (ParamKey.OTHER_COST, Alternative.INTERVENTION, ParamType.COST),
    "acei monitoring cost": (ParamKey.OTHER_COST, Alternative.COMPARATOR, ParamType.COST),
    "hf admission cost - base case": (ParamKey.EVENT_COST, Alternative.SHARED, ParamType.COST),
    "stable hfref utility": (ParamKey.BASELINE_UTILITY, Alternative.SHARED, ParamType.UTILITY),
    "qaly loss per hf admission": (ParamKey.EVENT_DISUTILITY, Alternative.SHARED, ParamType.DISUTILITY),
    "eligible hfref population": (ParamKey.ELIGIBLE_POPULATION, Alternative.SHARED, ParamType.COUNT),
    "low uptake": (ParamKey.UPTAKE_LOW, Alternative.SHARED, ParamType.RATE),
    "medium uptake": (ParamKey.UPTAKE_MEDIUM, Alternative.SHARED, ParamType.RATE),
    "high uptake": (ParamKey.UPTAKE_HIGH, Alternative.SHARED, ParamType.RATE),
}

# 01_INPUTS labels that map to EconomicModel scalars rather than parameters.
_SCALAR_MAP = {
    "time horizon": "horizon_years",
    "cost discount rate": "cost_discount_rate",
    "outcome discount rate": "outcome_discount_rate",
    "wtp threshold": "wtp_threshold",
}

# 04_PSA_INPUTS "Parameter" label -> (ParamKey, Alternative)
_PSA_MAP: dict[str, tuple[str, str]] = {
    "arni rehospitalization": (ParamKey.EVENT_PROBABILITY, Alternative.INTERVENTION),
    "acei rehospitalization": (ParamKey.EVENT_PROBABILITY, Alternative.COMPARATOR),
    "arni annual medicine cost": (ParamKey.DRUG_COST, Alternative.INTERVENTION),
    "acei annual medicine cost": (ParamKey.DRUG_COST, Alternative.COMPARATOR),
    "hf admission cost": (ParamKey.EVENT_COST, Alternative.SHARED),
    "stable utility": (ParamKey.BASELINE_UTILITY, Alternative.SHARED),
    "qaly loss/admission": (ParamKey.EVENT_DISUTILITY, Alternative.SHARED),
}

# 09_DATA_MAP_QC "Dashboard field" -> our metric name.
_QC_METRIC_MAP = {
    "total cost arni": "total_cost_intervention",
    "total cost acei": "total_cost_comparator",
    "total qaly arni": "total_qaly_intervention",
    "total qaly acei": "total_qaly_comparator",
    "incremental cost": "incremental_cost",
    "incremental effectiveness": "incremental_qaly",
    "icer": "icer",
    "inb": "inb",
    "bia low uptake": "bia_net_low",
    "bia medium uptake": "bia_net_medium",
    "bia high uptake": "bia_net_high",
    "psa probability cost-effective": "psa_prob_cost_effective",
}

_STATUS_MAP = {
    "observed": DataStatus.OBSERVED,
    "observed rwe": DataStatus.OBSERVED,
    "proxy": DataStatus.PROXY,
}


def _norm(text) -> str:
    """Lower-case, collapse whitespace, normalise en/em dashes to '-'."""
    if text is None:
        return ""
    s = str(text).replace("–", "-").replace("—", "-").replace("−", "-")
    return " ".join(s.split()).strip().lower()


def _dec(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def is_lecturer_workbook(sheetnames) -> bool:
    return LECTURER_SHEETS.issubset(set(sheetnames))


def _rows(ws) -> list[list]:
    return [[c.value for c in row] for row in ws.iter_rows()]


def parse_lecturer_workbook(file) -> dict:
    """Parse the lecturer's workbook into the same shape as `parse_workbook`."""
    wb = load_workbook(file, read_only=True, data_only=True)

    case_meta: dict[str, str] = {}
    scalars: dict[str, str] = {}
    params: dict[tuple, dict] = {}
    psa_config: dict[str, int] = {}

    # ── 01_INPUTS ────────────────────────────────────────────────────────
    for row in _rows(wb["01_INPUTS"]):
        if not row or row[0] is None:
            continue
        label = _norm(row[0])
        raw = row[1] if len(row) > 1 else None
        status = _norm(row[3]) if len(row) > 3 else ""
        source = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

        if label == "case id":
            case_meta["case_id"] = str(raw).strip() if raw is not None else ""
            continue
        if label == "psa simulations":
            psa_config["n_simulations"] = int(_dec(raw) or 1000)
            continue
        if label == "psa seed":
            psa_config["seed"] = int(_dec(raw) or 42)
            continue
        if label in _SCALAR_MAP:
            val = _dec(raw)
            if val is not None:
                scalars[_SCALAR_MAP[label]] = str(val)
            continue
        if label in _PARAM_MAP:
            key, alt, ptype = _PARAM_MAP[label]
            val = _dec(raw)
            if val is None:
                continue
            params[(key, alt)] = {
                "key": str(key),
                "alternative": str(alt),
                "year_index": None,
                "value": val,
                "param_type": str(ptype),
                "unit": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                "data_status": str(_STATUS_MAP.get(status, DataStatus.ASSUMPTION)),
                "source_reference": source,
                "distribution": "fixed",
                "dist_param1": None,
                "dist_param2": None,
            }

    # The engine's per-year BIA needs a base `uptake`; use the medium scenario.
    medium = params.get((ParamKey.UPTAKE_MEDIUM, Alternative.SHARED))
    if medium is not None:
        params[(ParamKey.UPTAKE, Alternative.SHARED)] = {
            **medium,
            "key": str(ParamKey.UPTAKE),
            "source_reference": "Diturunkan dari skenario uptake menengah (01_INPUTS)",
        }

    # ── 04_PSA_INPUTS: attach distributions ──────────────────────────────
    if "04_PSA_INPUTS" in wb.sheetnames:
        for row in _rows(wb["04_PSA_INPUTS"]):
            if not row or row[0] is None:
                continue
            label = _norm(row[0])
            if label not in _PSA_MAP:
                continue
            key, alt = _PSA_MAP[label]
            target = params.get((key, alt))
            if target is None:
                continue
            dist = _norm(row[1]) if len(row) > 1 else ""
            alpha = _dec(row[7]) if len(row) > 7 else None
            beta = _dec(row[8]) if len(row) > 8 else None
            if dist in {"beta", "gamma", "lognormal", "normal"} and alpha is not None:
                target["distribution"] = dist
                target["dist_param1"] = alpha
                target["dist_param2"] = beta

    # ── 09_DATA_MAP_QC: expected values + tolerances ─────────────────────
    expected_det: list[dict] = []
    expected_psa: list[dict] = []
    if "09_DATA_MAP_QC" in wb.sheetnames:
        for row in _rows(wb["09_DATA_MAP_QC"]):
            if not row or len(row) < 5 or row[1] is None:
                continue
            metric = _QC_METRIC_MAP.get(_norm(row[1]))
            if metric is None:
                continue
            expected = _dec(row[3])
            tolerance = _dec(row[4])
            if expected is None or tolerance is None:
                # e.g. QC12 whose expected cell is an uncomputed formula (#NAME?).
                continue
            item = {"metric": metric, "expected": expected, "tolerance": tolerance}
            if metric == "psa_prob_cost_effective":
                item.update(psa_config)
                expected_psa.append(item)
            else:
                expected_det.append(item)

    return {
        "missing_sheets": [],
        "format": "lecturer",
        "case_meta": case_meta,
        "model_scalars": scalars,
        "params": list(params.values()),
        "expected_deterministic": expected_det,
        "expected_psa": expected_psa,
        "psa_config": psa_config,
    }
