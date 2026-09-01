"""Build and parse the economic-validation workbook (Phase R6).

The lecturer's `DeciBridge_Economic_Validation_Model_ACEI_Dual.xlsx` was never
supplied, so we own the format: `build_workbook()` writes a workbook that encodes
the validation case (parameters + expected results), and `parse_workbook()` reads
one back. `validation_service.import_and_validate` maps a parsed workbook onto a
case, runs the engines, and compares actual vs expected.

Sheets:
    case_meta                    — field/value (case_id, title, drugs, indication)
    model_scalars                — field/value (horizon, discount rates, WTP, baseline)
    economic_model_params        — one row per parameter (key, alternative, value, ...)
    expected_deterministic_results — metric / expected / tolerance
    expected_psa_summary         — metric / expected / tolerance / n_simulations / seed
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook, load_workbook

from .validation_fixtures import (
    MODEL_SCALARS,
    VALIDATION_CASE_ID,
    VALIDATION_PARAMETERS,
)

PARAM_HEADER = [
    "key", "alternative", "year_index", "value", "param_type", "unit",
    "data_status", "distribution", "dist_param1", "dist_param2",
]

# Expected results for HF_ARNI_ACEI_001, mirroring the lecturer's workbook
# sheet 09_DATA_MAP_QC (QC01-QC11) including his tolerances.
EXPECTED_DETERMINISTIC = [
    ("total_cost_intervention", "18499451.85", "1"),
    ("total_cost_comparator", "5199411.1161", "1"),
    ("total_qaly_intervention", "0.655", "0.000001"),
    ("total_qaly_comparator", "0.62923", "0.000001"),
    ("incremental_cost", "13300040.7339", "1"),
    ("incremental_qaly", "0.02577", "0.000001"),
    ("icer", "516105577.5669392", "10"),
    ("inb", "-11109590.7339", "10"),
    ("bia_net_low", "133000407.339", "1"),
    ("bia_net_medium", "399001222.017", "1"),
    ("bia_net_high", "665002036.695", "1"),
]

# PSA expected (reproducible given the workbook's fixed seed + iteration count).
EXPECTED_PSA = [("psa_prob_cost_effective", "0.008", "0.001", "1000", "20260724")]


def build_workbook() -> Workbook:
    wb = Workbook()

    meta = wb.active
    meta.title = "case_meta"
    meta.append(["field", "value"])
    for field, value in [
        ("case_id", VALIDATION_CASE_ID),
        ("title", "ARNI vs ACEI pada pasien HFrEF"),
        ("intervention", "Sacubitril/Valsartan"),
        ("comparator", "Enalapril"),
        ("indication", "HFrEF"),
    ]:
        meta.append([field, value])

    scalars = wb.create_sheet("model_scalars")
    scalars.append(["field", "value"])
    for field, value in MODEL_SCALARS.items():
        scalars.append([field, str(value)])

    params = wb.create_sheet("economic_model_params")
    params.append(PARAM_HEADER)
    for spec in VALIDATION_PARAMETERS:
        params.append([
            str(spec["key"]),
            str(spec["alternative"]),
            "",  # year_index (blank = all years)
            str(spec["value"]),
            str(spec["param_type"]),
            spec.get("unit", ""),
            str(spec["data_status"]),
            spec.get("distribution", "fixed"),
            str(spec["dist_param1"]) if spec.get("dist_param1") is not None else "",
            str(spec["dist_param2"]) if spec.get("dist_param2") is not None else "",
        ])

    det = wb.create_sheet("expected_deterministic_results")
    det.append(["metric", "expected", "tolerance"])
    for row in EXPECTED_DETERMINISTIC:
        det.append(list(row))

    psa = wb.create_sheet("expected_psa_summary")
    psa.append(["metric", "expected", "tolerance", "n_simulations", "seed"])
    for row in EXPECTED_PSA:
        psa.append(list(row))

    return wb


def _rows(ws) -> list[list]:
    return [[c.value for c in row] for row in ws.iter_rows()]


def _field_value(ws) -> dict[str, str]:
    out: dict[str, str] = {}
    for row in _rows(ws)[1:]:  # skip header
        if row and row[0] is not None:
            out[str(row[0]).strip()] = "" if row[1] is None else str(row[1]).strip()
    return out


def parse_workbook(file) -> dict:
    """Read a workbook (path or file-like) into structured dicts."""
    wb = load_workbook(file, read_only=True, data_only=True)
    names = set(wb.sheetnames)
    required = {"case_meta", "model_scalars", "economic_model_params",
                "expected_deterministic_results", "expected_psa_summary"}
    missing_sheets = sorted(required - names)

    result: dict = {"missing_sheets": missing_sheets}
    if missing_sheets:
        return result

    result["case_meta"] = _field_value(wb["case_meta"])
    result["model_scalars"] = _field_value(wb["model_scalars"])

    param_rows = _rows(wb["economic_model_params"])
    header = [str(h).strip() if h is not None else "" for h in param_rows[0]]
    params = []
    for row in param_rows[1:]:
        if not row or row[0] is None:
            continue
        params.append({header[i]: row[i] for i in range(len(header))})
    result["params"] = params

    def expected(sheet, extra_cols):
        rows = _rows(wb[sheet])[1:]
        out = []
        for row in rows:
            if not row or row[0] is None:
                continue
            item = {"metric": str(row[0]).strip(), "expected": row[1], "tolerance": row[2]}
            for i, col in enumerate(extra_cols, start=3):
                item[col] = row[i] if i < len(row) else None
            out.append(item)
        return out

    result["expected_deterministic"] = expected("expected_deterministic_results", [])
    result["expected_psa"] = expected("expected_psa_summary", ["n_simulations", "seed"])
    return result


def _d(value) -> Decimal:
    return Decimal(str(value))
